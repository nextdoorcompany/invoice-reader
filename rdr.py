#! c:\venv\invenv\Scripts\python

import openpyxl
import sys
import collections
import pytest
import os
import datetime

Invoice = collections.namedtuple('Invoice', ['date', 'id', 'po', 'job', 'amount'])
Col_Map = collections.namedtuple('Col_Map', ['date', 'id', 'po', 'name', 'job', 'amount'])
xl_map = Col_Map(3, 4, 5, 6, 9, 10)
sheet_title = 'Sheet1'
start_row = 3
company = 'Next Door Distribution Company'
job_delim = ','

def get_invoices(filename):
	print('opening:  %s' % filename)
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_sheet_by_name(sheet_title)
	
	invoices = []
	for row in range(start_row, sheet.get_highest_row() + 1):
		if sheet.cell(row=row, column=xl_map.name).value == company:
			date = sheet.cell(row=row, column=xl_map.date).value.strftime('%m/%d/%Y')
			date_parts = [part.lstrip('0') for part in date.split('/')]
			date = '/'.join(date_parts)
			idn = sheet.cell(row=row, column=xl_map.id).value
			po = sheet.cell(row=row, column=xl_map.po).value
			jobs = sheet.cell(row=row, column=xl_map.job).value.split(job_delim)
			job = [j.strip(' ') for j in jobs]
			amount = sheet.cell(row=row, column=xl_map.amount).value
			amount = int(amount * 100)
			invoices.append(Invoice(date, idn, po, job, amount))

	return(invoices)

@pytest.fixture
def filename(request):
	fname = 'xtest.xlsx'

	def fin():
		os.remove(fname)
	request.addfinalizer(fin)
	return fname

def build_one_row_ss(filename, date, idn, po, name, job, amount):
	wb = openpyxl.Workbook()
	sheet = wb.get_active_sheet()
	sheet.title = sheet_title
	sheet.cell(row=start_row, column=xl_map.date).value = date
	sheet.cell(row=start_row, column=xl_map.id).value = idn
	sheet.cell(row=start_row, column=xl_map.po).value = po
	sheet.cell(row=start_row, column=xl_map.name).value = name
	sheet.cell(row=start_row, column=xl_map.job).value = job
	sheet.cell(row=start_row, column=xl_map.amount).value = amount

	wb.save(filename)

def test_rdr_basic(filename):
	build_one_row_ss(filename, datetime.datetime(2015, 9, 6), 12345, 'FL1027-010', company, '1027.R37', 456.22)
	result = get_invoices(filename)

	assert len(result) == 1
	assert result[0].date == '9/6/2015'
	assert result[0].id == 12345
	assert result[0].po == 'FL1027-010'
	assert result[0].job == ['1027.R37']
	assert result[0].amount == 45622

def test_rdr_not_ndd(filename):
	build_one_row_ss(filename, datetime.datetime(2015, 9, 6), 12345, 'FL1027-010', 'foo', '1027.R37', 456.22)
	result = get_invoices(filename)

	assert len(result) == 0

def test_rdr_zero_cents(filename):
	build_one_row_ss(filename, datetime.datetime(2015, 9, 6), 12345, 'FL1027-010', company, '1027.R37', 456.00)
	result = get_invoices(filename)

	assert len(result) == 1
	assert result[0].date == '9/6/2015'
	assert result[0].id == 12345
	assert result[0].po == 'FL1027-010'
	assert result[0].job == ['1027.R37']
	assert result[0].amount == 45600	

def test_rdr_over_thousand(filename):
	build_one_row_ss(filename, datetime.datetime(2015, 9, 6), 12345, 'FL1027-010', company, '1027.R37', 11456.23)
	result = get_invoices(filename)

	assert len(result) == 1
	assert result[0].date == '9/6/2015'
	assert result[0].id == 12345
	assert result[0].po == 'FL1027-010'
	assert result[0].job == ['1027.R37']
	assert result[0].amount == 1145623	

def test_rdr_two_jobs(filename):
	build_one_row_ss(filename, datetime.datetime(2015, 9, 6), 12345, 'FL1027-010', company, '1027.R37' + job_delim + '1027.R38', 11456.23)
	result = get_invoices(filename)

	assert len(result) == 1
	assert result[0].date == '9/6/2015'
	assert result[0].id == 12345
	assert result[0].po == 'FL1027-010'
	assert result[0].job == ['1027.R37', '1027.R38']
	assert result[0].amount == 1145623	

def test_rdr_two_jobs_extra_spaces(filename):
	build_one_row_ss(filename, datetime.datetime(2015, 9, 6), 12345, 'FL1027-010', company, '1027.R37' + job_delim + ' 1027.R38', 11456.23)
	result = get_invoices(filename)

	assert len(result) == 1
	assert result[0].date == '9/6/2015'
	assert result[0].id == 12345
	assert result[0].po == 'FL1027-010'
	assert result[0].job == ['1027.R37', '1027.R38']
	assert result[0].amount == 1145623	

